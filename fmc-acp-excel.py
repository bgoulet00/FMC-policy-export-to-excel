'''
This script will generate an excel xlsx file dump of a Cisco FMC access control policy
The output is meant to be a reference of the policy that appears similar in layout to the web GUI
The output is NOT suitable or intended to be a backup of the policy
Additionally, Users, Source Dynamic Attribute and Destination Dynamic Attribute are not accounted for but
updates to include those items should be trivial if you use those technologies
'''

# global variable BASE_URL will need to be updated with the url/IP of your FMC

# Developed and tested with the following environment
# OS: windows10
# Python: 3.11.5
# Target platform:  FMC 7.0.4
# Dependencies: protocols.csv input file with protocol number to name mappings. this is provided in the repo
# Limitations: 
#   - Users, Source Dynamic Attribute and Destination Dynamic Attribute are not accounted for
#   - FMC queries are limited to 1000 results. paging to support greater than 1000 has not been implemented
# Caveats: sometimes rules are added incorectly in FMC and appear to be part of a category when they are actually
#    'Undefined'.  this script will ignore those anomolies in an attempt to replicate the layout of the FMC GUI

import requests
from requests.auth import HTTPBasicAuth
import json
import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Disable SSL warnings
import urllib3
urllib3.disable_warnings()

# FMC URL/IP
BASE_URL = 'https://192.168.100.22'
# protocol info csv file with protocol number in column A and protocol name in column B. example A1 = 6, B1 = TCP
# the project repo provices a file based on iana data
protocol_file = 'protocols.csv'
# dictionary table to hold protocol number to name mappings as read from protocol_file. example {6:TCP,17:UDP,etc}
protocol_table = {}


# login to FMC and return the value of auth tokens and domain UUID from the response headers
# exit with an error message if a valid response is not received
def login():
    print('\n\nEnter FMC Credentials')
    user = input("USERNAME: ").strip()
    passwd = input("PASSWORD: ").strip()
    response = requests.post(
       BASE_URL + '/api/fmc_platform/v1/auth/generatetoken',
       auth=HTTPBasicAuth(username=user, password=passwd),
       headers={'content-type': 'application/json'},
       verify=False,
    )
    if response:
        return {'X-auth-access-token': response.headers['X-auth-access-token'], 
        'X-auth-refresh-token':response.headers['X-auth-refresh-token'],
        'DOMAIN_UUID':response.headers['DOMAIN_UUID']}
    else:
        sys.exit('Unable to connect to ' + BASE_URL + ' using supplied credentials')

#retrieve the list of access control policies in FMC
#limit is set to 1000, meaning only 1000 policies will be returned in a single query
#to support more than 1000 rules, this funtion would require modification to make successive calls, leveraging the 'pages' and 'offset'
#however, if you have more than 1000 polices it might be time for some cleanup or architecture redesign
def getPolicies(token, DUUID):
    response = requests.get(
       BASE_URL + '/api/fmc_config/v1/domain/' + DUUID + '/policy/accesspolicies?limit=1000',
       headers={'X-auth-access-token':token},
       verify=False,
    )
    raw = response.json()
    return raw

#for a given acess control policy ID, get all the rules using 'expanded' for full detail
#limit is set to 1000, meaning only 1000 rules will be returned in a single query
#to support more than 1000 rules, this funtion would require modification to make successive calls, leveraging the 'pages' and 'offset'
#however, if you have more than 1000 rules in a single ACP it might be time for some cleanup
def getRules(token, DUUID, acpID):
    response = requests.get(
       BASE_URL + '/api/fmc_config/v1/domain/' + DUUID + '/policy/accesspolicies/' + acpID + '/accessrules?limit=1000&expanded=true',
       headers={'X-auth-access-token':token},
       verify=False,
    )
    raw = response.json()
    return raw

#attempt to translate the protocol number to a name, else just return the number   
def protocolLookup(portObj):
    try:
        protocol = protocol_table[portObj['protocol']]
    except:
        protocol = portObj['protocol']
    return protocol

#populate protocol lookup table from csv file
def protocolTable():
    with open(protocol_file, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            protocol_table[row[0]] = row[1]

#create a string containing protocol, type, port, information
def interpretPorts(portObj):
    protocol = protocolLookup(portObj)
    if protocol == 'ICMP':
        portinfo = protocol + '/type ' + portObj['icmpType']
    elif protocol == 'TCP' or protocol == 'UDP':
        portinfo = protocol + '/' + portObj['port']
    else:
        portinfo = protocol
    return portinfo

#create excel file in similar format to FMC access control policy GUI 
def outputToExcel(filename, columns, policyRules):
    wb = Workbook() 
    ws = wb.active 
    
    #create header row in bold
    ws.append(columns)
    for i in range(len(columns)):
        ws.cell(row=1, column=i+1).font = Font(bold=True)

    #define colors for section and category separators
    darkGray = PatternFill(start_color='00BFBFBF',
                   end_color='00BFBFBF',
                   fill_type='solid')
    lightGray = PatternFill(start_color='00D9D9D9',
                   end_color='00D9D9D9',
                   fill_type='solid')
    
    #write each rule to a worksheet row
    #if a new section or category is encountered, create a separator row
    current_section = ''
    current_category = ''
    row = 2
    for rule in policyRules:
        if rule['section'] != current_section and rule['section'] != '--Undefined--':
            current_section = rule['section']
            separator = 'Section ' + current_section
            ws.cell(row=row, column=1).fill = darkGray
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            ws.cell(row=row, column=1).value = separator
        elif rule['category'] != current_category and rule['category'] != '--Undefined--':
            current_category = rule['category']
            separator = 'Category ' + current_category
            ws.cell(row=row, column=1).fill = lightGray
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            ws.cell(row=row, column=1).value = separator
        else:
            for i in range(len(columns)):
                ws.cell(row=row, column=i+1).alignment = Alignment(wrapText=True)
            ws.cell(row=row, column=1, value=rule['name'])
            ws.cell(row=row, column=2, value=rule['enabled'])
            ws.cell(row=row, column=3, value=rule['sourceZones'])
            ws.cell(row=row, column=4, value=rule['destinationZones'])
            ws.cell(row=row, column=5, value=rule['sourceNetworks'])
            ws.cell(row=row, column=6, value=rule['destinationNetworks'])
            ws.cell(row=row, column=7, value=rule['vlanTags'])
            ws.cell(row=row, column=8, value=rule['applications'])
            ws.cell(row=row, column=9, value=rule['sourcePorts'])
            ws.cell(row=row, column=10, value=rule['destinationPorts'])
            ws.cell(row=row, column=11, value=rule['urls'])
            ws.cell(row=row, column=12, value=rule['action'])
            ws.cell(row=row, column=13, value=rule['comments'])
        row = row + 1 

    # format column width to length of longest value plus padding
    # if the value is a multi-line string (contains \n) then format column
    # to the length of the longest piece in the multi-line string
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if '\n' in str(cell.value):
                pieces = str(cell.value).split('\n')
                for piece in pieces:
                   if len(piece) > max_length:
                        max_length = len(piece) 
            elif len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
 
    wb.save(filename)
    

def main():

    #login and retrieve token and DUUID
    result = login()
    token = result.get('X-auth-access-token')
    DUUID = result.get('DOMAIN_UUID')

    #list of dictionaries built from queried/manipulated data, each dict representing a complete policy rule
    policyRules = []
    #columns to be output to file, each column matching a dictionary key
    columns = ['name', 'enabled', 'sourceZones', 'destinationZones', 'sourceNetworks', 'destinationNetworks', 'vlanTags', 
                   'applications', 'sourcePorts', 'destinationPorts', 'urls', 'action', 'comments']
   
    #populate a dictionary with know protols per iana documentation
    protocolTable()

    #get the list of access control policies in FMC
    policies = getPolicies(token, DUUID)
    
    #prompt for input on which policy to export
    counter = 0
    print('Policies found')
    for item in policies['items']:
        counter = counter +1
        print('[',counter,']',item['name'])
    entry = int(input('Enter the number of the policy you want to export: '))

    #get the rules associated with the policy, rules being a list[] of rules in the policy
    rules = getRules(token, DUUID, policies['items'][entry -1]['id'])

    #iterate through the rules, extracting the fields/data and copy them to a new list of dicts
    #FMC 'mostly' does not return keys with empty values so this section will normalize all keys
    #empty or non-existing keys will be created with the value 'Any'
    #keys that happen to be lists will be converted to a string of \n separated values for multiline output in a cell
    temp_list = []
    for rule in rules['items']:
        new_rule = {}
        new_rule['enabled'] = rule['enabled']
        new_rule['name'] = rule['name']
        new_rule['action'] = rule['action']
        new_rule['section'] = rule['metadata']['section']
        new_rule['category'] = rule['metadata']['category']
        if 'commentHistoryList' in rule.keys():
            temp_list.clear()
            for item in rule['commentHistoryList']:
                temp_list.append(item['comment'])
            new_rule['comments'] = '\n'.join(temp_list)
        else: new_rule['comments'] = ''
        if 'sourceZones' in rule.keys(): 
            temp_list.clear()
            for item in rule['sourceZones']['objects']:
                temp_list.append(item['name'])
            new_rule['sourceZones'] = '\n'.join(temp_list)
        else: new_rule['sourceZones'] = 'Any'
        if 'destinationZones' in rule.keys(): 
            temp_list.clear()
            for item in rule['destinationZones']['objects']:
                temp_list.append(item['name'])
            new_rule['destinationZones'] = '\n'.join(temp_list)
        else: new_rule['destinationZones'] = 'Any'
        if 'sourceNetworks' in rule.keys(): 
            temp_list.clear()
            if 'objects' in rule['sourceNetworks']:
                for item in rule['sourceNetworks']['objects']:
                    temp_list.append(item['name'])
            if 'literals' in rule['sourceNetworks']:
                for item in rule['sourceNetworks']['literals']:
                    temp_list.append(item['value'])
            new_rule['sourceNetworks'] = '\n'.join(temp_list)
        else: new_rule['sourceNetworks'] = 'Any'
        if 'destinationNetworks' in rule.keys(): 
            temp_list.clear()
            if 'objects' in rule['destinationNetworks']:
                for item in rule['destinationNetworks']['objects']:
                    temp_list.append(item['name'])
            if 'literals' in rule['destinationNetworks']:
                for item in rule['destinationNetworks']['literals']:
                    temp_list.append(item['value'])
            new_rule['destinationNetworks'] = '\n'.join(temp_list)
        else: new_rule['destinationNetworks'] = 'Any'
        if 'vlanTags' in rule.keys(): #this key is an anomoly in that FMC returns it even if it is empty
            temp_list.clear()
            if 'objects' in rule['vlanTags']:
                for item in rule['vlanTags']['objects']:
                    temp_list.append(item['name'])
            if 'literals' in rule['vlanTags']:
                for item in rule['vlanTags']['literals']:
                    vlan_range = str(item['startTag']) + '-' + str(item['endTag'])
                    temp_list.append(vlan_range)
            else: temp_list.append('Any')
            new_rule['vlanTags'] = '\n'.join(temp_list)
        else: new_rule['vlanTags'] = 'Any'
        if 'applications' in rule.keys():
            temp_list.clear()
            for item in rule['applications']['applications']:
                    temp_list.append(item['name'])
            new_rule['applications'] = '\n'.join(temp_list)
        else: new_rule['applications'] = 'Any'
        if 'sourcePorts' in rule.keys():
            temp_list.clear()
            if 'objects' in rule['sourcePorts']:
                for item in rule['sourcePorts']['objects']:
                    temp_list.append(item['name'])
            if 'literals' in rule['sourcePorts']:
                for item in rule['sourcePorts']['literals']:
                    literal = interpretPorts(item)
                    temp_list.append(literal)
            new_rule['sourcePorts'] = '\n'.join(temp_list)
        else: new_rule['sourcePorts'] = 'Any'
        if 'destinationPorts' in rule.keys():
            temp_list.clear()
            if 'objects' in rule['destinationPorts']:
                for item in rule['destinationPorts']['objects']:
                    temp_list.append(item['name'])
            if 'literals' in rule['destinationPorts']:
                for item in rule['destinationPorts']['literals']:
                    literal = interpretPorts(item)
                    temp_list.append(literal)
            new_rule['destinationPorts'] = '\n'.join(temp_list)
        else: new_rule['destinationPorts'] = 'Any'
        if 'urls' in rule.keys():
            temp_list.clear()
            if 'urlCategoriesWithReputation' in rule['urls']:
                for item in rule['urls']['urlCategoriesWithReputation']:
                    url = item['category']['name'] + '/' + item['reputation']
                    temp_list.append(url)
            if 'literals' in rule['urls']:
                for item in rule['urls']['literals']:
                    temp_list.append(item['url'])
            new_rule['urls'] = '\n'.join(temp_list)
        else: new_rule['urls'] = 'Any'
        policyRules.append(new_rule)


    #output to excel
    excel_file = 'FMC-ACP-' + policies['items'][entry -1]['name'] + '.xlsx'
    outputToExcel(excel_file, columns, policyRules)
    print('\nXLSX output for access control policy', policies['items'][entry -1]['name'], 'complete')

if __name__ == "__main__":
    main()

